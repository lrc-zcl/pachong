import os
import re
import time
from openpyxl import Workbook, load_workbook
from lxml import etree
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver import ChromeOptions

"""
爬取高考帮网站所有高职院校的数据信息
注意： 又是或获取不到数据 上下滑动滑动滚轮就行了
"""
chrome_option = Options()
chrome_option.add_argument("--headless")  # 禁用弹框
chrome_option.add_argument("--disable-gpu")
chrome_option.add_argument("--disable-images")  # 禁用图片

# 实现规避检测
option = ChromeOptions()
option.add_experimental_option("excludeSwitches", ["enable-automation"])


class pro_info():
    """
    获取 学校名、学校具体位置、学校性质、学校类型
    """

    def __init__(self):
        self.bro = webdriver.Chrome(chrome_options=chrome_option, options=option)
        self.bro.get("https://www.gaokao.cn/")
        self.update_page_result()

    def update_page_result(self):
        """
        更新网页信息，获取到最新的网页数据
        :return:
        """
        self.page_result = self.bro.page_source
        self.tree = etree.HTML(self.page_result)

    def hualun(self):

        # 获取当前页面的滚动位置和页面总高度
        scroll_height = self.bro.execute_script("return document.body.scrollHeight;")
        scroll_position = self.bro.execute_script("return window.scrollY + window.innerHeight;")

        # 判断是否在页面最上面
        if scroll_position <= 0:
            print("At the top of the page. Scrolling to the bottom.")
            # 滚动到页面底部
            self.bro.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        elif scroll_position >= scroll_height:
            print("At the bottom of the page. Scrolling to the top.")
            # 滚动到页面顶部
            self.bro.execute_script("window.scrollTo(0, 0);")
        else:
            self.bro.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            self.bro.execute_script("window.scrollTo(0, 0);")
            print("In the middle of the page. every once.")

        # 等待一段时间以便页面响应滚动
        time.sleep(0.5)

    def huoqu_daxue_counts(self):
        """
        获取一共有多少页大学
        :return:
        """
        self.hualun()
        self.update_page_result()
        ele_data = self.tree.xpath('//div[@class="school-search_schoolItem__3q7R2"]')
        return len(ele_data)

    def remove_kuohao(self, s):
        """
        去除字符串中括号的内容
        :return:
        """
        return re.sub(r'（[^）]*）|\([^)]*\)', '', s).strip()

    def monidenglu(self):
        """
        模拟登录
        :return:
        """
        self.hualun()
        denglu = self.bro.find_element(by=By.XPATH,
                                       value='//span[@class="login-btn-float_loginText__9o_uo"]')
        denglu.click()
        self.hualun()

        mimadenglu = self.bro.find_element(by=By.XPATH,
                                           value='//div[@class="login-popup_loginMode__1mHjy"]/div[2]/div[2]')
        mimadenglu.click()
        self.hualun()

        zhanghao = self.bro.find_element(by=By.XPATH,
                                         value='//div[@class="login-popup_inputItem__29c36"]/input')
        zhanghao.send_keys("13137741301")

        mima = self.bro.find_element(by=By.XPATH,
                                     value='//input[@type="password"]')
        mima.send_keys('Lrc130530')
        self.hualun()

        yuedu = self.bro.find_element(by=By.XPATH,
                                      value='//span[@class="ant-checkbox"]/input')
        yuedu.click()
        self.hualun()

        final_denglu = self.bro.find_element(by=By.XPATH,
                                             value='//div[@class="login-popup_loginBtn__3buCc "]')
        final_denglu.click()

    def chadaxue(self):
        """
        模拟登录完成，点击查大学
        :return:
        """
        # 首页->点击查大学
        self.hualun()
        chadaxue = self.bro.find_element(by=By.XPATH,
                                         value='//div[@class="main-nav_navBox__MY9It"]/div[3]/div[1]')
        chadaxue.click()

    def dianji_diqu(self, index):
        """
        点击查看大学
        :param index: 某一个地区对应的位置索引
        :return:
        """
        # //*[@id="root"]/div/div[1]/div/div/div[2]/div/div[3]/div/div[1]/div[1]/div/div[4]/div[2]
        self.hualun()
        base_xpath = '//div[@class="school-search_schoolListMain__B9yLk"]/div[4]/div[2]'
        xuexiao_click_xpath = base_xpath + f'/span[{index}]'
        xuexiao_click_weizhi = self.bro.find_element(by=By.XPATH, value=xuexiao_click_xpath)
        xuexiao_click_weizhi.click()  # 点击

    def dianji_zhuanke(self):
        """
        点击专科 高职
        :return:
        """

        self.hualun()
        zhuanke_xpath = '//div[@class="school-search_schoolListMain__B9yLk"][1]/div[6]/div[2]/span[3]'
        zhuanke_data = self.bro.find_element(by=By.XPATH, value=zhuanke_xpath)
        zhuanke_data.click()

    def huoqu_daxue_yeshu(self):
        """
        点击专科之后，获取大学的页数
        :return:
        """
        self.hualun()
        self.update_page_result()
        data = len(self.tree.xpath(
            '///ul[@class="ant-pagination "]/li'))
        if data == 0:
            page_counts = 1
        else:
            page_counts = data - 2

        return page_counts

    def dianji_mouyige_daxue(self, index):
        """
        根据当前页中 大学的个数量
        :param index:
        :return:
        """
        self.hualun()
        base_xpath = '//div[@class="school-search_schoolItem__3q7R2"]'
        xuexiao_click_xpath = base_xpath + f'[{index}]' + "/div[2]/div[1]/h3/em"

        dianjimouyige = self.bro.find_element(by=By.XPATH, value=xuexiao_click_xpath)
        dianjimouyige.click()

    def dianji_kaishe_zhuanye(self):
        self.hualun()
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.common.by import By

        # 等待元素可点击
        element = WebDriverWait(self.bro, 100).until(
            EC.element_to_be_clickable((By.XPATH,
                                        '//*[@id="root"]/div/div[1]/div/div/div[1]/div[2]/div[3]/div/div[4]/div/div[2]/div[4] | //*[@id="root"]/div/div[1]/div/div/div[1]/div[3]/div[3]/div/div[4]/div/div[2]/div[4]'))
        )
        from selenium.webdriver import ActionChains
        actions = ActionChains(self.bro)
        actions.move_to_element_with_offset(element, 5, 5).click().perform()

    def huoqu_xuexiao_name(self):
        self.hualun()
        self.hualun()
        self.update_page_result()
        xuexiaoname = self.tree.xpath('//div[@class="school-tab_info__3x6H6"]/div[1]/div[1]//text()')[0]
        location = self.tree.xpath('//div[@class="school-tab_info__3x6H6"]/div[1]/div[2]//text()')[0]
        data = self.tree.xpath('//div[@class="school-tab_info__3x6H6"]/div[2]/div')
        # print(location)

        if len(data) >= 3:
            gongbanmingban = self.tree.xpath('//div[@class="school-tab_info__3x6H6"]/div[2]/div[3]//text()')[0]
            xuexiaoleixing = self.tree.xpath('//div[@class="school-tab_info__3x6H6"]/div[2]/div[2]//text()')[0]
        else:
            gongbanmingban = self.tree.xpath('//div[@class="school-tab_info__3x6H6"]/div[2]/div[2]//text()')[0]
            xuexiaoleixing = '不明确'
        return [xuexiaoname, location, gongbanmingban, xuexiaoleixing]

    def huoqu_zhuanye_yeshu(self):
        self.hualun()
        self.update_page_result()
        data = len(self.tree.xpath('//*[@id="setUpMajorEle"]/div[3]/div/ul/li'))
        if data == 0:
            zhuanyeyeshu = 1
        else:
            zhuanyeyeshu = data - 2

        return zhuanyeyeshu

    def huoqu_mouyiye_zhuanye_data(self, zhuanye_ele_counts):
        """
        利用当前页专业ele list
        :param zhuanye_ele_counts:
        :return:
        """
        zhuanyename_list = []
        xuekemenlei_list = []
        zhuanyeleibie_list = []
        self.hualun()
        self.update_page_result()
        for ele in zhuanye_ele_counts:
            zhuanyename = ele.xpath('./td[1]//text()')[0]
            xuekemenlei = ele.xpath('./td[3]//text()')[0]
            zhuanyeleibie = ele.xpath('./td[4]//text()')[0]
            zhuanyename_list.append(zhuanyename)
            xuekemenlei_list.append(xuekemenlei)
            zhuanyeleibie_list.append(zhuanyeleibie)

        result = list(zip(zhuanyename_list, xuekemenlei_list, zhuanyeleibie_list))
        return result

    def xlsxfile(self):
        """
        创建工作表并写入表头
        :return:
        """
        diqu_ele_data = self.tree.xpath(
            '//*[@id="root"]/div/div[1]/div/div/div[2]/div/div[3]/div/div[1]/div[1]/div/div[4]/div[2]/span')
        file_path = "../data/all_data.xlsx"
        # headers = ["school", "location", "nature", "type", "pro", "subject_categories", "pro_category", "course",
        # "remarks"]
        headers = ["学校", "位置", "公办/民办", "类型", "专业名称", "学科门类", "专业类别", "主修课程",
                   "备注"]

        if not os.path.exists(file_path):
            self.wb = Workbook()
            self.wb.save(file_path)  # 创建文件

            self.workbook = load_workbook(file_path)
            for i in diqu_ele_data[1:]:  # 第一个是全部 从北京开始
                diqudata = i.xpath('.//text()')[0]
                self.workbook.create_sheet(title=diqudata)  # 创建工作表
                if 'Sheet' in self.workbook.sheetnames:
                    del self.workbook['Sheet']
                for index, headerdata in enumerate(headers, start=1):
                    self.workbook[diqudata].cell(1, index, headerdata)  # 创建表头
                    self.workbook.save(file_path)
            print("表格文件创建完成！！")
        else:
            self.workbook = load_workbook(file_path)

    def write_xslx_data(self, diquname, hangindex, data):
        file_path = "../data/all_data.xlsx"
        for index, signal_data, in enumerate(data, start=1):
            self.workbook[diquname].cell(hangindex, index, signal_data)
        self.workbook.save(file_path)

    def genju_zhuanye_chaxiangqing(self, zhuanye_name, kaishi_xpath):
        """
        diyici : '//*[@id="root"]/div/div[1]/div/div/div[1]/div[2]/div[1]/div[6]/div/div[4]/div'
        houmin = '//*[@id="root"]/div/div[1]/div/div/div[2]/div/div[1]/div[6]/div/div[4]/div'
        :param zhuanye_name:
        :param kaishi_xpath:
        :return:
        """
        # self.hualun()
        chazhuanye = self.bro.find_element(by=By.XPATH,
                                           value=kaishi_xpath)

        chazhuanye.click()
        # print(1)

        self.hualun()  # 网络好的情况下 可以试试注释

        dianjikaozhi = self.bro.find_element(by=By.XPATH,
                                             value='//div[@class="filter-compents_itemBox__1Q3EV"]/span[2]')

        dianjikaozhi.click()
        # print(2)

        # self.hualun()

        zhuanye_sousuo = self.bro.find_element(by=By.XPATH, value='//input[@placeholder="输入专业名称"]')
        zhuanye_sousuo.clear()
        zhuanye_sousuo.send_keys(zhuanye_name)  # 根据专业名称搜索专业

        button_sousuo = self.bro.find_element(by=By.XPATH,
                                              value='//*[@id="root"]/div/div[1]/div/div/div[2]/div/div[3]/div/div[1]/div/div[1]/div[1]/div[1]/button')
        button_sousuo.click()  # 点击搜索
        # print(3)

        self.hualun()

        # 搜索后会新建窗口
        dianji_zhiqian = self.bro.current_window_handle

        sousuohou_de_diyige = self.bro.find_element(by=By.XPATH,
                                                    value='//div[@class="major-list_sonList__29bFH"]')
        sousuohou_de_diyige.click()
        # print(4)

        all_winds = self.bro.window_handles
        for wins in all_winds:
            if wins != dianji_zhiqian and wins != self.first_wins:
                self.bro.switch_to.window(wins)  # 切换到新的窗口

        self.hualun()
        self.update_page_result()  # 更新页面数据
        kecheng = self.tree.xpath(
            '//*[@id="root"]/div/div[1]/div/div/div[2]/div/div[3]/div/div/div[3]/div/div/div[1]/div[3]/div[2]//text()')
        beizhuxinxi = self.tree.xpath(
            '//*[@id="root"]/div/div[1]/div/div/div[2]/div/div[3]/div/div/div[3]/div/div/div[1]/div[3]/div[1]//text()')
        self.bro.close()  # 关闭新建的窗口
        self.bro.switch_to.window(self.second_wins)  # 关闭窗口后立即跳转到第二个窗口

        return kecheng, beizhuxinxi

    def main(self):
        """
        因为过一段时间 就会登录过期 所以需要重新登录一下
        :return:
        """
        self.monidenglu()
        self.hualun()
        self.chadaxue()
        self.hualun()
        self.update_page_result()
        self.xlsxfile()
        for i in range(23, 25):
            hang_sum_data = 0
            self.dianji_diqu(index=i)
            self.update_page_result()
            diquname = \
                self.tree.xpath(f'//div[@class="school-search_schoolListMain__B9yLk"]/div[4]/div[2]/span[{i}]//text()')[
                    0]
            self.dianji_zhuanke()
            daxueyeshu = self.huoqu_daxue_yeshu()  # 这里还是第一个窗口

            for daxue_page in range(daxueyeshu):
                daxue_counts = self.huoqu_daxue_counts()  # 每一页的大学数量

                for i_data in range(daxue_counts):  # 遍历每一个大学
                    self.first_wins = self.bro.current_window_handle
                    self.dianji_mouyige_daxue(i_data + 1)  # 现在就开始了第二个窗口
                    all_window = self.bro.window_handles  # 获取所有窗口

                    for window in all_window:
                        if window != self.first_wins:
                            self.second_wins = window
                            self.bro.switch_to.window(window)  # 切换到新窗口
                            break
                    xuexiao_base_info_list = self.huoqu_xuexiao_name()
                    self.dianji_kaishe_zhuanye()

                    zhuanye_page_counts = self.huoqu_zhuanye_yeshu()  # 获取专业的页数
                    for y in range(zhuanye_page_counts):
                        self.chazhuanye_xpath = '//div[@class="main-nav_navBox__MY9It"]/div[4]/div[1]'
                        self.hualun()
                        self.update_page_result()
                        zhuanye_ele_counts = self.tree.xpath('//*[@id="setUpMajorEle"]/div[2]/div/table/tbody/tr')
                        result = self.huoqu_mouyiye_zhuanye_data(zhuanye_ele_counts)

                        signal_page_all_zhuanye_result = [list(zip_data) for zip_data in result]  # 每一页的所有专业

                        # 根据当前页所有的专业 获取专业所学课程
                        final_all_info_result_2d = []
                        for index, zhuanye_data in enumerate(signal_page_all_zhuanye_result):  # 开始对每一个专业继续搜索 获取详细信息
                            wanzheng_zhuanye_name = zhuanye_data[0]
                            final_zhuanye_name = self.remove_kuohao(s=wanzheng_zhuanye_name)
                            try:
                                kecheng, beizhu = self.genju_zhuanye_chaxiangqing(final_zhuanye_name,
                                                                                  kaishi_xpath=self.chazhuanye_xpath)
                            except:
                                print(f"{xuexiao_base_info_list} + {final_zhuanye_name} + 没找到")
                                kecheng = ["没找到"]
                                beizhu = ["没找到"]
                            # self.chazhuanye_xpath = '//*[@id="root"]/div/div[1]/div/div/div[2]/div/div[1]/div[6]/div/div[4]/div'
                            kecheng = " ".join(kecheng)
                            final_beizhu = wanzheng_zhuanye_name + " " + " ".join(beizhu)
                            signal_page_all_zhuanye_result[index].extend([kecheng, final_beizhu])  # 将课程和备注信息添加到 专业的后面
                            final_all_info_result = xuexiao_base_info_list + signal_page_all_zhuanye_result[index]
                            final_all_info_result_2d.append(final_all_info_result)
                            print(final_all_info_result)

                            self.write_xslx_data(diquname=diquname, hangindex=hang_sum_data + 2,
                                                 data=final_all_info_result)
                            hang_sum_data = hang_sum_data + 1
                        print("这一页已加载完成！！")

                        # 执行完一页专业后 返回第二个窗口首页 点击开设专业  并点击 下一页
                        self.bro.switch_to.window(self.second_wins)
                        self.hualun()
                        while True:
                            try:
                                self.bro.back()
                                if self.bro.find_element(by=By.XPATH,
                                                         value='//div[@class="school-tab_tabNavs__1wdWg"]/div[4]'):
                                    print("开设专业已经找到！！")
                                    self.dianji_kaishe_zhuanye()
                                    self.hualun()
                                    if y != zhuanye_page_counts - 1:
                                        for page_i in range(y + 1):  # 点击几次下一页
                                            zhuanye_xiayiye = self.bro.find_element(by=By.XPATH,
                                                                                    value=f'//*[@id="setUpMajorEle"]/div[3]/div/ul/li[{zhuanye_page_counts + 2}]')
                                            zhuanye_xiayiye.click()  # 点击下一页
                                    break
                            except Exception as e:
                                # print(e)
                                print("有问题，没找到开设专业，继续找！！")
                                continue

                    self.bro.close()  # 第一个学校已完成  切换到第一窗口 点击下一个学校
                    self.bro.switch_to.window(self.first_wins)
                    self.hualun()

                # 第一页的大学已经获取完全  现在回到首页 开始获取第二页的大学
                # 为了避免登陆过期  现在每一页 新的大学就开始重新登录一下
                if daxue_page != daxueyeshu - 1:
                    self.__init__()
                    self.hualun()
                    self.hualun()
                    self.hualun()
                    self.monidenglu()
                    self.hualun()
                    self.chadaxue()
                    self.hualun()
                    self.update_page_result()
                    # self.xlsxfile()
                    self.dianji_diqu(index=i)
                    self.update_page_result()
                    diquname = self.tree.xpath(
                        f'//div[@class="school-search_schoolListMain__B9yLk"]/div[4]/div[2]/span[{i}]//text()')[0]

                    self.dianji_zhuanke()
                    self.hualun()

                    # 点击大学的下一页
                    daxue_xiayiye = self.bro.find_element(by=By.XPATH,
                                                          value=f'//ul[@class="ant-pagination "]/li[{daxueyeshu + 2}]')
                    for data in range(daxue_page + 1):
                        daxue_xiayiye.click()
                        self.hualun()

                # if daxue_page != daxueyeshu - 1:
                #     # self.bro.close()
                #     self.bro.switch_to.window(self.first_wins)
                #     self.hualun()
                #     daxue_xiayiye = self.bro.find_element(by=By.XPATH,
                #                                           value=f'//*[@id="root"]/div/div[1]/div/div/div[2]/div/div[3]/div/div[1]/div[2]/div/div[2]/div/ul/li[{daxueyeshu + 2}]')
                #     daxue_xiayiye.click()  # 点击下一页大学

            # 切换到第一个窗口 点击下一个地区
            # self.bro.close()
            # self.bro.switch_to.window(self.first_wins)
            # self.hualun()
            print(f"{diquname}已完成！！")


if __name__ == '__main__':
    base = pro_info()
    base.main()
